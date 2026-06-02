"""报告生成与导出。

将 PipelineResult 转换为可导出的报告，支持 JSON/Excel/CSV/HTML 格式。
"""
import csv
import io
import json
import os
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from app.engine.pipeline import PipelineResult, FileProcessResult, StageTrace
from app.engine.ipe import Person
from app.engine.risk import risk_distribution

_tz = timezone(timedelta(hours=8))


def _now() -> str:
    return datetime.now(_tz).strftime("%Y-%m-%d %H:%M:%S")


def _ts() -> str:
    return datetime.now(_tz).strftime("%Y%m%d_%H%M%S")


@dataclass
class ReportSummary:
    total_files: int = 0
    total_persons: int = 0
    total_entities: int = 0
    error_files: int = 0
    by_type: dict = field(default_factory=dict)


def build_summary(result: PipelineResult) -> ReportSummary:
    s = ReportSummary()
    s.total_files = len(result.files)
    s.total_persons = result.total_persons
    s.error_files = sum(1 for f in result.files if f.error)
    for f in result.files:
        if f.extraction_result:
            for e in f.extraction_result.entities:
                s.total_entities += 1
                s.by_type[e.type] = s.by_type.get(e.type, 0) + 1
    return s


def export_json(result: PipelineResult) -> str:
    """导出为 JSON 字符串。"""
    persons_dicts = [p.to_dict() for p in result.all_persons]
    from app.engine.risk import add_risk_to_person_dict
    for pd in persons_dicts:
        add_risk_to_person_dict(pd)

    data = {
        "generated_at": _now(),
        "summary": {
            "total_files": len(result.files),
            "total_persons": result.total_persons,
            "risk_distribution": risk_distribution(persons_dicts),
        },
        "files": [],
        "persons": persons_dicts,
    }
    for f in result.files:
        entry = {
            "path": f.file_path,
            "ext": f.ext,
            "size_mb": f.size_mb,
            "error": f.error,
            "entities": [],
            "persons": len(f.ipe_result.persons) if f.ipe_result else 0,
            "traces": [],
        }
        if f.extraction_result:
            entry["entities"] = [{"type": e.type, "value": e.value, "confidence": e.confidence}
                                 for e in f.extraction_result.entities]
        if f.masked_report:
            entry["masked"] = f.masked_report.get("masked_text", "")
        if f.traces:
            entry["traces"] = [{"stage": t.stage, "status": t.status,
                                "fields_count": t.fields_count,
                                "error": t.error, "duration_ms": round(t.duration_ms, 1)}
                               for t in f.traces]
        data["files"].append(entry)
    return json.dumps(data, ensure_ascii=False, indent=2)


def export_csv(result: PipelineResult) -> str:
    """导出为 CSV 字符串（人员维度）。"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["姓名", "身份证号", "手机号", "邮箱", "银行卡号", "地址", "微信",
                     "生日", "工号", "性别", "车牌号", "护照号", "置信度"])
    for p in result.all_persons:
        d = p.to_dict()
        writer.writerow([
            d.get("name", ""),
            d.get("id_card", ""),
            d.get("phone", ""),
            d.get("email", ""),
            d.get("bank_card", ""),
            d.get("address", ""),
            d.get("wechat", ""),
            d.get("birthday", ""),
            d.get("job_no", ""),
            d.get("gender", ""),
            d.get("plate_no", ""),
            d.get("passport", ""),
            d.get("confidence", ""),
        ])
    return output.getvalue()


def export_excel(result: PipelineResult) -> bytes:
    """导出为 Excel 二进制数据。"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl 未安装，无法导出 Excel")

    wb = Workbook()

    # Sheet 1: 人员报表
    ws = wb.active
    ws.title = "数据主体"
    headers = ["姓名", "身份证号", "手机号", "邮箱", "银行卡号", "地址", "微信",
               "生日", "工号", "性别", "车牌号", "护照号", "置信度"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row, p in enumerate(result.all_persons, 2):
        d = p.to_dict()
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=d.get(h, ""))

    # Sheet 2: 文件概览
    ws2 = wb.create_sheet("文件概览")
    file_headers = ["文件路径", "后缀", "大小(MB)", "实体数", "状态"]
    for col, h in enumerate(file_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for row, f in enumerate(result.files, 2):
        ws2.cell(row=row, column=1, value=f.file_path)
        ws2.cell(row=row, column=2, value=f.ext)
        ws2.cell(row=row, column=3, value=f.size_mb)
        count = len(f.extraction_result.entities) if f.extraction_result else 0
        ws2.cell(row=row, column=4, value=count)
        ws2.cell(row=row, column=5, value=f.error or "OK")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_html(result: PipelineResult) -> str:
    """导出为独立 HTML 页面。"""
    persons_rows = ""
    for p in result.all_persons:
        d = p.to_dict()
        persons_rows += "<tr>"
        for key in ["name", "id_card", "phone", "email", "bank_card", "address",
                     "wechat", "birthday", "job_no", "gender", "plate_no", "passport"]:
            persons_rows += f"<td>{d.get(key, '') or ''}</td>"
        persons_rows += f"<td>{d.get('confidence', '')}</td></tr>\n"

    files_rows = ""
    for f in result.files:
        count = len(f.extraction_result.entities) if f.extraction_result else 0
        files_rows += f"<tr><td>{f.file_path}</td><td>{f.ext}</td><td>{f.size_mb}</td>"
        files_rows += f"<td>{count}</td><td>{f.error or 'OK'}</td></tr>\n"

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>DocScanX 扫描报告</title>
<style>
body {{ font-family: 'Microsoft YaHei', sans-serif; max-width: 1200px; margin: 0 auto; padding: 20px; color: #333; }}
h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
.summary {{ display: flex; gap: 24px; margin: 20px 0; }}
.summary-item {{ background: #f8f9fa; padding: 16px 24px; border-radius: 8px; text-align: center; }}
.summary-item .value {{ font-size: 28px; font-weight: bold; color: #3498db; }}
.summary-item .label {{ font-size: 13px; color: #888; margin-top: 4px; }}
table {{ width: 100%; border-collapse: collapse; margin: 16px 0; font-size: 13px; }}
th {{ background: #3498db; color: #fff; padding: 10px 8px; text-align: left; }}
td {{ padding: 8px; border-bottom: 1px solid #eee; }}
tr:hover {{ background: #f5f8ff; }}
.footer {{ margin-top: 30px; color: #999; font-size: 12px; text-align: center; }}
</style>
</head>
<body>
<h1>DocScanX 扫描报告</h1>
<div class="summary">
<div class="summary-item"><div class="value">{len(result.files)}</div><div class="label">扫描文件</div></div>
<div class="summary-item"><div class="value">{result.total_persons}</div><div class="label">识别身份</div></div>
</div>
<h2>数据主体</h2>
<table>
<thead><tr><th>姓名</th><th>身份证号</th><th>手机号</th><th>邮箱</th><th>银行卡</th><th>地址</th><th>微信</th><th>生日</th><th>工号</th><th>性别</th><th>车牌</th><th>护照</th><th>置信度</th></tr></thead>
<tbody>{persons_rows}</tbody>
</table>
<h2>文件概览</h2>
<table>
<thead><tr><th>文件路径</th><th>后缀</th><th>大小(MB)</th><th>实体数</th><th>状态</th></tr></thead>
<tbody>{files_rows}</tbody>
</table>
<div class="footer">DocScanX 生成于 {_now()}</div>
</body>
</html>"""


def save_report(result: PipelineResult, output_dir: str, report_id: str = None,
                 task_name: str = "") -> str:
    """将报告保存到输出目录，返回报告 ID。"""
    rid = report_id or _ts()
    report_dir = os.path.join(output_dir, rid)
    os.makedirs(report_dir, exist_ok=True)

    # JSON
    with open(os.path.join(report_dir, "report.json"), "w", encoding="utf-8") as f:
        f.write(export_json(result))

    # CSV
    with open(os.path.join(report_dir, "report.csv"), "w", encoding="utf-8-sig") as f:
        f.write(export_csv(result))

    # Excel
    xlsx_data = export_excel(result)
    with open(os.path.join(report_dir, "report.xlsx"), "wb") as f:
        f.write(xlsx_data)

    # HTML
    with open(os.path.join(report_dir, "report.html"), "w", encoding="utf-8") as f:
        f.write(export_html(result))

    # 元数据
    meta = {"id": rid, "name": task_name or rid, "generated_at": _now(),
            "total_files": len(result.files),
            "total_persons": result.total_persons,
            "error_files": sum(1 for f in result.files if f.error),
            "total_entities": sum(len(f.extraction_result.entities) for f in result.files if f.extraction_result)}
    with open(os.path.join(report_dir, "meta.json"), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False)

    return rid


def list_reports(output_dir: str) -> List[dict]:
    """列出输出目录中的所有报告。"""
    reports = []
    if not os.path.isdir(output_dir):
        return reports
    for name in sorted(os.listdir(output_dir), reverse=True):
        p = os.path.join(output_dir, name)
        if not os.path.isdir(p):
            continue
        meta_file = os.path.join(p, "meta.json")
        if os.path.exists(meta_file):
            with open(meta_file, "r", encoding="utf-8") as f:
                reports.append(json.load(f))
    return reports
